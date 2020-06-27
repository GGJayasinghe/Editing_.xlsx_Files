import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']

    # 5 is the starting row of the table
    # 17 is the end row of the table

    for row in range(5, 17):
        cell = sheet.cell(row, 3)

        # this is the calculation we should perform with data set
        corrected_value = cell.value * 0.9

        # 12 represent the no of columns in the data set
        corrected_value_cell = sheet.cell(row, 12)
        corrected_value_cell.value = corrected_value

    # draw the table with the data in col 12 from 5th row to 16th row s
    values = Reference(sheet, min_row=5, max_row=16, min_col=12, max_col=12)
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'n3')

    wb.save(filename)


file_name = input("Enter the file name: ")
file_name = file_name + ".xlsx"
process_workbook(file_name)
